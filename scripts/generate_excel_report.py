import os
import json
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font

parser = argparse.ArgumentParser()

parser.add_argument("--reports-dir", required=True)
parser.add_argument("--app-name", required=True)
parser.add_argument("--run-id", required=True)
parser.add_argument("--output", required=True)

args = parser.parse_args()

reports_dir = args.reports_dir
app_name = args.app_name
run_id = args.run_id
output_file = args.output

summary_file = os.path.join(reports_dir, "scan-summary.json")

summary = {}

if os.path.exists(summary_file):

    with open(summary_file, "r", encoding="utf-8") as f:
        summary = json.load(f)

else:

    summary = {
        "critical_count": 0,
        "high_count": 0,
        "medium_count": 0,
        "low_count": 0,
        "secrets_count": 0
    }

wb = Workbook()

ws = wb.active
ws.title = "Security Report"

title = "Dynamic DevSecOps Security Report"

ws["A1"] = title
ws["A1"].font = Font(bold=True, size=16)

ws["A3"] = "Application Name"
ws["B3"] = app_name

ws["A4"] = "Pipeline Run ID"
ws["B4"] = run_id

ws["A6"] = "Severity"
ws["B6"] = "Count"

headers = ["Severity", "Count"]

for cell in ws[6]:
    cell.font = Font(bold=True)

data = [
    ["Critical", summary.get("critical_count", 0)],
    ["High", summary.get("high_count", 0)],
    ["Medium", summary.get("medium_count", 0)],
    ["Low", summary.get("low_count", 0)],
    ["Secrets", summary.get("secrets_count", 0)],
]

row = 7

for item in data:

    ws.cell(row=row, column=1, value=item[0])
    ws.cell(row=row, column=2, value=item[1])

    row += 1

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 20

wb.save(output_file)

print("==================================================")
print("EXCEL REPORT GENERATED")
print("==================================================")
print(f"Application : {app_name}")
print(f"Run ID      : {run_id}")
print(f"Output File : {output_file}")
print("==================================================")